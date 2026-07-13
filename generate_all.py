import json
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
import qrcode
from pathlib import Path
from PIL import Image, ImageDraw, ImageFont

# =========================
# SETTINGS
# =========================
EXCEL_FILE = "staff.xlsx"
SHEET_NAME = "staff"

BASE_URL = "https://enamecard-one.vercel.app"

PEOPLE_JSON = "people.json"
WHATSAPP_XLSX = "whatsapp_urls.xlsx"
REGISTRATION_NO = "201001040122 (924047-P)"
HQ_ADDRESS = "16 & 16a, Jalan Firma 2/1, Kawasan Perindustrian Tebrau 1, 81100 Johor Bahru, Johor, Malaysia."
HQ_TEL = "07-361 6565"
HQ_FAX = "07-361 5373"

CONTACT_DIR = Path("contacts")
QR_DIR = Path("qr")
DESIGNED_QR_DIR = Path("designed_qr")

CONTACT_DIR.mkdir(exist_ok=True)
QR_DIR.mkdir(exist_ok=True)
DESIGNED_QR_DIR.mkdir(exist_ok=True)

# =========================
# READ EXCEL
# =========================
wb = load_workbook(EXCEL_FILE, data_only=True)
ws = wb[SHEET_NAME]

headers = [cell.value for cell in ws[1]]
people = {}
whatsapp_rows = []

FONT_PATH = Path(__file__).parent / "fonts" / "EthnocentricRg.ttf"
LOGO_PATH = Path(__file__).with_name("SENG LI + 3icon.png")
RIGHT_ICON_PATH = Path(__file__).with_name("3icon.png")
LEFT_LINE_PATH = Path(__file__).with_name("right line.png")
WINDOWS_FONT_DIR = Path("C:/Windows/Fonts")


def load_font(size, fallback="arial.ttf"):
    try:
        if FONT_PATH.exists():
            return ImageFont.truetype(str(FONT_PATH), size)
        return ImageFont.truetype(fallback, size)
    except Exception:
        return ImageFont.load_default()


def load_windows_font(filename, size, fallback="arial.ttf"):
    font_path = WINDOWS_FONT_DIR / filename
    try:
        if font_path.exists():
            return ImageFont.truetype(str(font_path), size)
        return ImageFont.truetype(fallback, size)
    except Exception:
        return ImageFont.load_default()


def fit_font(text, size, max_width, fallback="arial.ttf", min_size=22):
    measure = ImageDraw.Draw(Image.new("RGB", (1, 1)))
    while size >= min_size:
        font = load_font(size, fallback)
        bbox = measure.textbbox((0, 0), text, font=font)
        if bbox[2] - bbox[0] <= max_width:
            return font
        size -= 2
    return load_font(min_size, fallback)


def draw_left_aligned_text(draw, x, y, text, font, fill, anchor="la", stroke_width=0, stroke_fill=None):
    """
    Draw text so the visible left edge starts exactly at x.
    This fixes small font side-bearing offsets, so Name/Title/Company/HQ line up visually.
    """
    bbox = draw.textbbox((0, 0), text, font=font, stroke_width=stroke_width)
    adjusted_x = x - bbox[0]
    draw.text(
        (adjusted_x, y),
        text,
        fill=fill,
        font=font,
        anchor=anchor,
        stroke_width=stroke_width,
        stroke_fill=stroke_fill if stroke_fill else fill
    )


def render_tracking_text(text, font, fill, tracking=-3, space_adjust=-8, stroke_width=0, stroke_fill=None):
    """
    Render text with custom letter spacing.
    Negative tracking reduces the wide default spacing of display fonts like Ethnocentric.
    The returned image is cropped, so the visible left edge can align exactly.
    """
    measure = ImageDraw.Draw(Image.new("RGBA", (1, 1), (0, 0, 0, 0)))

    widths = []
    heights = []
    left_offsets = []
    right_offsets = []

    for ch in text:
        bbox = measure.textbbox((0, 0), ch, font=font, stroke_width=stroke_width)
        left_offsets.append(bbox[0])
        right_offsets.append(bbox[2])
        widths.append(bbox[2] - bbox[0])
        heights.append(bbox[3] - bbox[1])

    if not widths:
        return Image.new("RGBA", (1, 1), (0, 0, 0, 0))

    # Estimate canvas width. Spaces get extra negative adjustment because the font's default word gap is wide.
    total_w = 0
    for i, ch in enumerate(text):
        total_w += widths[i]
        if i < len(text) - 1:
            total_w += tracking
            if ch == " " or text[i + 1] == " ":
                total_w += space_adjust

    bbox_all = measure.textbbox((0, 0), text, font=font, stroke_width=stroke_width)
    total_h = bbox_all[3] - bbox_all[1]
    pad = 10

    img = Image.new("RGBA", (max(1, int(total_w + pad * 2)), max(1, int(total_h + pad * 2))), (0, 0, 0, 0))
    d = ImageDraw.Draw(img)

    cursor_x = pad
    baseline_y = pad - bbox_all[1]
    for i, ch in enumerate(text):
        ch_bbox = measure.textbbox((0, 0), ch, font=font, stroke_width=stroke_width)
        d.text(
            (cursor_x - ch_bbox[0], baseline_y),
            ch,
            font=font,
            fill=fill,
            stroke_width=stroke_width,
            stroke_fill=stroke_fill if stroke_fill else fill
        )
        cursor_x += widths[i]
        if i < len(text) - 1:
            cursor_x += tracking
            if ch == " " or text[i + 1] == " ":
                cursor_x += space_adjust

    crop = img.getbbox()
    if crop:
        img = img.crop(crop)
    return img


def paste_tracking_name(card, x, center_y, text, font, fill, tracking=-3, space_adjust=-8, stroke_width=1, stroke_fill="#F2C517"):
    """
    Paste the name with corrected letter spacing and exact visible-left alignment.
    No font auto-shrink is used.
    """
    name_img = render_tracking_text(
        text=text,
        font=font,
        fill=fill,
        tracking=tracking,
        space_adjust=space_adjust,
        stroke_width=stroke_width,
        stroke_fill=stroke_fill
    )
    card.paste(name_img, (int(x), int(center_y - name_img.height / 2)), name_img)




def paste_tracking_text_left_top(card, x, y, text, font, fill, tracking=0.5, space_adjust=0, stroke_width=0, stroke_fill=None):
    """
    Paste text with custom letter spacing from a fixed top-left visual position.
    Useful for company name tracking without changing font size.
    """
    text_img = render_tracking_text(
        text=text,
        font=font,
        fill=fill,
        tracking=tracking,
        space_adjust=space_adjust,
        stroke_width=stroke_width,
        stroke_fill=stroke_fill if stroke_fill else fill
    )
    card.paste(text_img, (int(x), int(y)), text_img)


def apply_qr_yellow_transparency(qr_rgb, opacity=1.0):
    """
    Convert QR to brand-yellow modules with transparent black background.
    opacity=0.90 means 10% transparent.
    """
    qr_rgba = qr_rgb.convert("RGBA")
    pixels = qr_rgba.load()
    alpha = int(255 * opacity)

    for y in range(qr_rgba.height):
        for x in range(qr_rgba.width):
            r, g, b, a = pixels[x, y]
            # QR modules are brand yellow; background is black.
            if r > 128 or g > 128 or b > 128:
                pixels[x, y] = (242, 197, 23, alpha)
            else:
                pixels[x, y] = (0, 0, 0, 0)

    return qr_rgba

for row_values in ws.iter_rows(min_row=2, values_only=True):
    row = dict(zip(headers, row_values))

    person_id = str(row.get("id") or "").strip()

    if person_id == "":
        continue

    name = str(row["name"]).strip()
    title = str(row["title"]).strip()
    company = str(row["company"]).strip()
    phone = str(row["phone"]).strip()
    email = str(row["email"]).strip()
    whatsapp = str(row["whatsapp"]).strip()
    website = str(row["website"]).strip()
    photo = str(row["photo"]).strip()

    vcf_path = f"contacts/{person_id}.vcf"
    url = f"{BASE_URL}/?id={person_id}"

    people[person_id] = {
        "name": name,
        "title": title,
        "company": company,
        "phone": phone,
        "email": email,
        "whatsapp": whatsapp,
        "website": website,
        "photo": photo,
        "vcf": vcf_path,
        "url": url
    }
    whatsapp_rows.append((name, url))

    # =========================
    # GENERATE VCF
    # =========================
    vcf_content = f"""BEGIN:VCARD
VERSION:3.0
FN:{name}
ORG:{company}
TITLE:{title}
TEL;TYPE=WORK,CELL:+{phone}
EMAIL:{email}
URL:{website}
END:VCARD
"""

    with open(CONTACT_DIR / f"{person_id}.vcf", "w", encoding="utf-8") as f:
        f.write(vcf_content)

    # =========================
    # GENERATE QR
    # =========================
    qr = qrcode.QRCode(
        version=1,
        box_size=12,
        border=2
    )
    qr.add_data(url)
    qr.make(fit=True)

    qr_img = qr.make_image(
        fill_color="#F2C517",
        back_color="black"
    ).convert("RGB")

    qr_img.save(QR_DIR / f"{person_id}.png")

    # =========================
    # GENERATE DESIGNED QR CARD
    # =========================
    card_width = 1300
    card_height = 800
    safe_area = 30

    card = Image.new("RGB", (card_width, card_height), "#000000")
    draw = ImageDraw.Draw(card)

    # =========================
    # LEFT BRAND IMAGE
    # =========================
    left_content_shift = 30
    left_line_width = 48
    try:
        left_line = Image.open(LEFT_LINE_PATH).convert("RGBA")
        left_line_height = card_height
        left_line = left_line.resize((left_line_width, left_line_height), Image.LANCZOS)
        card.paste(left_line, (left_content_shift, 0), left_line)
    except Exception as e:
        print(f"Left line image error: {e}")

    # =========================
    # LOGO
    # =========================
    logo_y = safe_area
    logo_height = 0
    try:
        logo = Image.open(LOGO_PATH).convert("RGBA")

        logo_max_width = 650
        logo_max_height = 205

        logo.thumbnail(
            (logo_max_width, logo_max_height),
            Image.LANCZOS
        )

        logo_x = card_width - safe_area - logo.width
        logo_height = logo.height
        card.paste(
            logo,
            (logo_x, logo_y),
            logo
        )

    except Exception as e:
        print(f"Logo error: {e}")

    # =========================
    # RIGHT BOTTOM ICON
    # =========================
    try:
        right_icon = Image.open(RIGHT_ICON_PATH).convert("RGBA")
        right_icon.thumbnail((260, 95), Image.LANCZOS)
        right_icon_x = card_width - safe_area - right_icon.width
        right_icon_y = card_height - safe_area - right_icon.height
        card.paste(right_icon, (right_icon_x, right_icon_y), right_icon)
    except Exception as e:
        print(f"Right icon error: {e}")

    # =========================
    # FONTS
    # =========================
    font_name = load_font(40, "arialbd.ttf")
    font_title = load_windows_font("seguib.ttf", 32, "arialbd.ttf")
    try:
        font_company = load_windows_font("seguib.ttf", 25, "arialbd.ttf")
        font_company_small = ImageFont.truetype("seguisb.ttf", 18)
        font_hq = load_windows_font("seguisb.ttf", 20, "arialbd.ttf")
    except Exception:
        font_company = ImageFont.load_default()
        font_company_small = ImageFont.load_default()
        font_hq = ImageFont.load_default()

    # =========================
    # CONTENT POSITION
    # =========================
    content_center_y = 540

    # =========================
    # COMPANY INFO LAYOUT
    # =========================
    # Company name + company no. are grouped together with HQ info,
    # but separated into 2 adjustable sub-groups:
    # 1) Company name + company code
    # 2) Address + Tel + Fax
    company_info_y = 590

    # You can adjust these 3 values independently:
    company_code_gap = 30        # distance between company name and company code
    group_gap = 40               # distance between company code group and address/tel/fax group
    contact_line_gap = 30        # same line spacing for address, tel, fax group

    # =========================
    # QR
    # =========================
    qr_size = 150

    qr_resized = qr_img.resize(
        (qr_size, qr_size),
        Image.LANCZOS
    )
    qr_resized = apply_qr_yellow_transparency(qr_resized, opacity=1.0)

    # =========================
    # LEFT TEXT
    # =========================
    left_x = left_content_shift + left_line_width + safe_area + 45
    qr_x = left_x

    line_name = 70
    line_title = 57

    total_text_height = line_name + line_title

    text_margin_top = 20

    text_start_y = (
        content_center_y
        - total_text_height / 2
        + text_margin_top
        - 70
    )

    qr_y = logo_y + int((logo_height - qr_size) / 2) if logo_height else safe_area
    card.paste(
        qr_resized,
        (qr_x, qr_y),
        qr_resized
    )

    paste_tracking_name(
        card=card,
        x=left_x,
        center_y=text_start_y + 15,
        text=name,
        font=font_name,
        fill="#F2C517",
        tracking=1,
        space_adjust=0,
        stroke_width=1,
        stroke_fill="#F2C517"
    )

    draw_left_aligned_text(
        draw,
        left_x,
        text_start_y + line_name,
        title,
        fill="white",
        font=font_title,
        anchor="lm"
    )


    # =========================
    # COMPANY INFO
    # =========================
    # Company name + company no. are now grouped together with HQ info,
    # but separated into 2 adjustable sub-groups:
    # 1) Company name + company code
    # 2) Address + Tel + Fax

    label_hq = "HQ: "
    label_tel = "Tel: "
    label_fax = "Fax: "
    address_line_1 = "16 & 16A, Jalan Firma 2/1, Kawasan Perindustrian Tebrau 1,"
    address_line_2 = "81100 Johor Bahru, Johor, Malaysia."

    y = company_info_y

    # ----- Group 1: Company name + company code -----
    # Company Name with slight letter spacing (+0.5)
    paste_tracking_text_left_top(
        card=card,
        x=left_x,
        y=y,
        text=company,
        font=font_company,
        fill="#FFFFFF",
        tracking=0.5,
        space_adjust=0,
        stroke_width=0
    )

    y += company_code_gap
    draw_left_aligned_text(
        draw, left_x, y,
        REGISTRATION_NO,
        fill="#8E8E8E",
        font=font_company_small,
        anchor="la"
    )

    # ----- Gap between group 1 and group 2 -----
    y += group_gap

    # ----- Group 2: Address + Tel + Fax -----
    draw_left_aligned_text(draw, left_x, y, label_hq, fill="#F2C517", font=font_hq)
    hq_bbox = draw.textbbox((0, 0), label_hq, font=font_hq)
    hq_label_width = hq_bbox[2] - hq_bbox[0]
    draw.text((left_x + hq_label_width, y), address_line_1, fill="#DDDDDD", font=font_hq)

    y += contact_line_gap
    draw_left_aligned_text(draw, left_x, y, address_line_2, fill="#DDDDDD", font=font_hq)

    y += contact_line_gap
    draw_left_aligned_text(draw, left_x, y, label_tel, fill="#F2C517", font=font_hq)
    tel_bbox = draw.textbbox((0, 0), label_tel, font=font_hq)
    tel_label_width = tel_bbox[2] - tel_bbox[0]
    draw.text((left_x + tel_label_width, y), HQ_TEL, fill="#DDDDDD", font=font_hq)

    y += contact_line_gap
    draw_left_aligned_text(draw, left_x, y, label_fax, fill="#F2C517", font=font_hq)
    fax_bbox = draw.textbbox((0, 0), label_fax, font=font_hq)
    fax_label_width = fax_bbox[2] - fax_bbox[0]
    draw.text((left_x + fax_label_width, y), HQ_FAX, fill="#DDDDDD", font=font_hq)

    # =========================
    # SAVE DESIGNED CARD
    # =========================
    card.save(
        DESIGNED_QR_DIR / f"{person_id}_enamecard.png"
    )

# =========================
# EXPORT people.json
# =========================
with open(PEOPLE_JSON, "w", encoding="utf-8") as f:
    json.dump(people, f, ensure_ascii=False, indent=2)

# =========================
# EXPORT WhatsApp URL Excel
# =========================
url_wb = Workbook()
url_ws = url_wb.active
url_ws.title = "WhatsApp URLs"
url_ws.append(["name", "url"])

header_fill = PatternFill("solid", fgColor="000000")
header_font = Font(color="F2C517", bold=True)

for cell in url_ws[1]:
    cell.fill = header_fill
    cell.font = header_font

for row_index, (name, url) in enumerate(whatsapp_rows, start=2):
    url_ws.cell(row=row_index, column=1, value=name)
    url_cell = url_ws.cell(row=row_index, column=2, value=url)
    url_cell.hyperlink = url
    url_cell.style = "Hyperlink"

url_ws.column_dimensions["A"].width = 26
url_ws.column_dimensions["B"].width = 52
url_ws.freeze_panes = "A2"
url_wb.save(WHATSAPP_XLSX)

print("Done.")
print("Generated:")
print("- people.json")
print("- contacts/*.vcf")
print("- qr/*.png")
print("- designed_qr/*.png")
print("- whatsapp_urls.xlsx")
